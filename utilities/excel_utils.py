from openpyxl import load_workbook


class ExcelUtils:

    @staticmethod
    def get_data(path, sheet):

        workbook = load_workbook(path)

        worksheet = workbook[sheet]

        data = []

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            data.append(row)

        return data